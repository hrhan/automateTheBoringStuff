#! python3

import openpyxl, argparse
from openpyxl.styles import Font, PatternFill, colors
from os import getcwd

parser = argparse.ArgumentParser()
parser.add_argument('number', type = int)
args = parser.parse_args()

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Multiplication'

marginFont = Font(size = 13, bold = True)
marginFill = PatternFill("solid", fgColor = 'C5D9F1')

for i in range(1, args.number+1):
	sheet.cell(row = 1, column = i+1).value = i
	sheet.cell(row = 1, column = i+1).font = marginFont
	sheet.cell(row = 1, column = i+1).fill = marginFill
	sheet.cell(row = i+1, column = 1).value = i
	sheet.cell(row = i+1, column = 1).font = marginFont
	sheet.cell(row = i+1, column = 1).fill = marginFill
	
for row in range(2, args.number+2):
	for col in range(2, args.number+2):
		sheet.cell(row = row, column = col).value = (row-1)*(col-1)	

excelFile = 'multiplication_{}.xlsx'.format(args.number)		
wb.save(excelFile)
print('Saved {} to {}'.format(excelFile, getcwd()))
	