#! python3

import openpyxl, os, argparse
from sys import exit

def select_sheet_by_index(workbook, index):
	try:
		sheet = workbook.get_sheet_by_name(workbook.sheetnames[index])
		print('{} selected'.format(sheet.title))
		return sheet
	except:
		print('Invalid worksheet selected. Exiting program..')
		exit()

if __name__ == '__main__':
	parser = argparse.ArgumentParser()
	parser.add_argument('start', type = int)
	parser.add_argument('count', type = int)
	parser.add_argument('file')
	parser.add_argument('--w', '-worksheet', type=int)
	args = parser.parse_args()

	if not os.path.exists(os.path.abspath(args.file)):
		print('Unable to find {}. Exiting program..'.format(args.file))
		exit()
		
	wb = openpyxl.load_workbook(args.file)
	if args.w:
		sheetIdx = args.w
	else:
		sheetIdx = wb.index(wb.active)	
	sheet = select_sheet_by_index(wb, sheetIdx)
	maxRow = sheet.max_row
	maxCol = sheet.max_column

	for row in reversed(range(args.start + args.count, maxRow + args.count + 1)):
		for col in range(1, maxCol + 1):
			sheet.cell(row = row, column = col).value = sheet.cell(row = row - args.count, column = col).value
			
	for row in range(args.start, args.start + args.count):
		for col in range(1, maxCol + 1):
			sheet.cell(row = row, column = col).value = None
		
	wb.save(args.file)
	print('{} blank rows starting from row {} added to {}'.format(args.count, args.start, args.file))