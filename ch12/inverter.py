#! python3

import openpyxl, os, argparse
from sys import exit
from blankRowInserter import select_sheet_by_index

parser = argparse.ArgumentParser()
parser.add_argument('file')
parser.add_argument('--w', '-worksheet', type = int)
args = parser.parse_args()

if not os.path.exists(os.path.abspath(args.file)):
	print('Unable to find {}. Exiting the program..'.format(args.file))
	exit()
	
wb = openpyxl.load_workbook(args.file)
if args.w:
	sheetIdx = args.w
else:
	sheetIdx = wb.index(wb.active)	
sheet = select_sheet_by_index(wb, sheetIdx)		
invertedSheet = wb.create_sheet(title = sheet.title + "_inverted", index = wb.index(sheet) + 1)

for row in range(1, sheet.max_row + 1):
	for col in range(1, sheet.max_column + 1):
		invertedSheet.cell(row = col, column = row).value = sheet.cell(row=row, column=col).value

wb.save(args.file)
print('Saved inveted result to {} in {}'.format(invertedSheet.title, args.file))