#! python3

import openpyxl, os, argparse
from sys import exit
from blankRowInserter import select_sheet_by_index

parser = argparse.ArgumentParser()
parser.add_argument('files', nargs = '+')
parser.add_argument('mode', choices = ['text2spread', 'spread2text'])
parser.add_argument('--w', '-worksheet', type = int)
args = parser.parse_args()

if args.mode == 'text2spread':
	wb = openpyxl.Workbook()
	sheet = wb.active

	for idx, file in enumerate(args.files):
		if not os.path.exists(os.path.abspath(file)):
			print('Unable to find {}. Moving on to the next file..'.format(file))
			continue
		
		with open(file) as textFile:
			print('Working on {}..'.format(file))
			sheet.cell(row = 1, column = idx + 1).value = file
			for lineNum, line in enumerate(textFile):
				sheet.cell(row = lineNum + 2, column = idx + 1).value = line.strip()
		print('Finished copying contents from {}'.format(file))
	wbName = '-'.join([os.path.splitext(file)[0] for file in args.files]) + '.xlsx'
	wb.save(wbName)
	print('Text files saved to {}'.format(wbName))
	
elif args.mode == 'spread2text':
	for file in args.files:
		if not os.path.exists(os.path.abspath(file)):
			print('Unable to find {}. Moving on to the next file..'.format(file))
			continue
		
		print('Working on {}..'.format(file))
		textName = os.path.splitext(file)[0] + '.txt'
		
		with open(textName, 'w') as textFile:
			wb = openpyxl.load_workbook(file)
			for worksheet in wb.sheetnames:
				sheet = wb[worksheet]
				if sheet.dimensions=='A1:A1':
					continue
				else:
					print('Copying contents from {}..'.format(worksheet))
					for col in range(1, sheet.max_column + 1):
						for row in range(1, sheet.max_row + 1):
							textFile.write(str(sheet.cell(row = row, column = col).value) + '\n')
				textFile.write('\n')
		wb.close()
		print('Spreadsheets saved to {}'.format(textName))